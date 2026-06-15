from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook


SUBJECT_COLS = {
    4: "基础",
    5: "写作",
    6: "口语",
    7: "阅读",
    8: "听力",
}

SUBJECT_MAP = {
    "writing": "写作",
    "speaking": "口语",
    "reading": "阅读",
    "listening": "听力",
    "grammar": "基础",
    "写作": "写作",
    "口语": "口语",
    "阅读": "阅读",
    "听力": "听力",
    "基础": "基础",
    "语法": "基础",
    "W": "写作",
    "S": "口语",
    "R": "阅读",
    "L": "听力",
}

DATE_RE = re.compile(
    r"(?P<m>\d{1,2})[./](?P<d>\d{1,2})\s+"
    r"(?P<sh>\d{1,2})[:.](?P<sm>\d{2})\s*-\s*"
    r"(?P<eh>\d{1,2})[:.](?P<em>\d{2})"
    r"(?:\s*(?P<tail>.*))?"
)

SUBJECT_LINE_RE = re.compile(
    r"(?i)\b(writing|speaking|reading|listening|grammar)\b|写作|口语|阅读|听力|基础|语法"
)

BAD_TEACHER_TOKENS = {
    "余",
    "剩",
    "结课",
    "沟通",
    "搬家",
    "回复",
    "开始排",
    "基础",
    "语法",
}


@dataclass
class ParsedCourse:
    source_row: int
    student_name: str
    course_type: str
    teacher_name: str
    course_date: str
    start_time: str
    end_time: str
    duration_hours: float


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def strip_hours(value: str) -> str:
    return re.sub(r"\d+(?:\.\d+)?\s*[Hh小时]*", "", value).strip()


def normalize_teacher_name(value: str) -> str:
    name = clean(value).strip(" \t\r\n，,;；：:")
    if not name:
        return ""
    if any(token in name for token in BAD_TEACHER_TOKENS):
        return ""
    if len(name) > 18:
        return ""
    if not re.search(r"[A-Za-z\u4e00-\u9fff]", name):
        return ""
    if re.fullmatch(r"[WSRL]\d*(?:\.\d+)?", name, re.I):
        return ""
    return name


def teacher_candidates_from_cell(value) -> list[str]:
    teachers: list[str] = []
    for raw_line in clean(value).splitlines():
        candidate = normalize_teacher_name(strip_hours(raw_line))
        if candidate:
            teachers.append(candidate)
    return teachers


def subject_from_text(line: str) -> str | None:
    lowered = line.lower()
    for key in ["writing", "speaking", "reading", "listening", "grammar"]:
        if key in lowered:
            return SUBJECT_MAP[key]
    for key in ["写作", "口语", "阅读", "听力", "语法", "基础"]:
        if key in line:
            return SUBJECT_MAP[key]
    return None


def infer_subject_from_code(value) -> str | None:
    text = clean(value).upper()
    subjects: list[str] = []
    if "W" in text:
        subjects.append("写作")
    if "S" in text:
        subjects.append("口语")
    if "R" in text:
        subjects.append("阅读")
    if "L" in text:
        subjects.append("听力")
    return subjects[0] if len(subjects) == 1 else None


def calculate_duration(start_time: str, end_time: str) -> float:
    start_dt = datetime.strptime(start_time, "%H:%M")
    end_dt = datetime.strptime(end_time, "%H:%M")
    return round((end_dt - start_dt).total_seconds() / 3600, 2)


def now_iso() -> str:
    return datetime.now(UTC).replace(tzinfo=None).isoformat()


def collect_known_teachers(ws) -> set[str]:
    known: set[str] = set()
    for row in range(2, ws.max_row + 1):
        for col in SUBJECT_COLS:
            known.update(teacher_candidates_from_cell(ws.cell(row, col).value))

    for row in range(2, ws.max_row + 1):
        schedule_text = clean(ws.cell(row, 11).value)
        for raw_line in schedule_text.splitlines():
            line = clean(raw_line)
            if DATE_RE.search(line) or not SUBJECT_LINE_RE.search(line):
                continue
            rest = SUBJECT_LINE_RE.sub("", line)
            candidate = normalize_teacher_name(strip_hours(rest))
            if candidate:
                known.add(candidate)

    return known


def parse_teacher_map(ws, row: int, known_teachers: set[str]) -> dict[str, str]:
    teacher_map: dict[str, str] = {}
    for col, subject in SUBJECT_COLS.items():
        candidates = [
            teacher
            for teacher in teacher_candidates_from_cell(ws.cell(row, col).value)
            if teacher in known_teachers
        ]
        if candidates:
            teacher_map[subject] = candidates[-1]
    return teacher_map


def parse_tail(tail: str, known_teachers: set[str]) -> tuple[str | None, str | None, float | None]:
    text = re.sub(r"[（(].*?[）)]", "", clean(tail)).strip()

    duration = None
    duration_match = re.search(r"\b(\d+(?:\.\d+)?)\b", text)
    if duration_match:
        duration = float(duration_match.group(1))

    subject = None
    for token in re.findall(r"\b[WSRL]\b", text, re.I):
        subject = SUBJECT_MAP[token.upper()]

    teacher = None
    for known_teacher in sorted(known_teachers, key=len, reverse=True):
        if known_teacher and known_teacher in text:
            teacher = known_teacher
            break

    return subject, teacher, duration


def parse_workbook(xlsx_path: Path, sheet_name: str, year: int, month: int) -> tuple[list[dict], set[str], list[ParsedCourse]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    known_teachers = collect_known_teachers(ws)

    students: list[dict] = []
    courses: list[ParsedCourse] = []

    for row in range(2, ws.max_row + 1):
        student_name = clean(ws.cell(row, 2).value)
        if not student_name:
            continue

        end_date = clean(ws.cell(row, 12).value)
        remark = clean(ws.cell(row, 14).value)
        schedule_text = clean(ws.cell(row, 11).value)
        status = "在读"
        if "暂停" in remark or "暂不排课" in schedule_text:
            status = "暂停"
        elif end_date or "已结课" in schedule_text or "结课" in schedule_text:
            status = "结课"

        note_parts = [
            f"来源：7月课时.xlsx 第 {row} 行",
            f"总课时数：{clean(ws.cell(row, 3).value) or '-'}",
            f"7月完成课时：{clean(ws.cell(row, 13).value) or '-'}",
        ]
        if end_date:
            note_parts.append(f"结课日期：{end_date}")
        if remark:
            note_parts.append(f"备注：{remark}")

        students.append(
            {
                "name": student_name,
                "status": status,
                "note": "；".join(note_parts),
            }
        )

        teacher_map = parse_teacher_map(ws, row, known_teachers)
        current_subject: str | None = None
        current_teacher: str | None = None
        fallback_subject = next(iter(teacher_map.keys()), None) or infer_subject_from_code(ws.cell(row, 3).value)

        for raw_line in schedule_text.splitlines():
            line = clean(raw_line)
            if not line:
                continue

            date_match = DATE_RE.search(line)
            if date_match:
                parsed_month = int(date_match.group("m"))
                if parsed_month != month:
                    continue

                day = int(date_match.group("d"))
                start_time = f"{int(date_match.group('sh')):02d}:{date_match.group('sm')}"
                end_hour = int(date_match.group("eh"))
                if end_hour >= 24:
                    continue
                end_time = f"{end_hour:02d}:{date_match.group('em')}"
                computed_duration = calculate_duration(start_time, end_time)
                if computed_duration <= 0:
                    continue

                tail_subject, tail_teacher, tail_duration = parse_tail(date_match.group("tail") or "", known_teachers)
                course_type = tail_subject or current_subject or fallback_subject or "自定义"
                teacher_name = tail_teacher or current_teacher or teacher_map.get(course_type) or "未分配老师"

                courses.append(
                    ParsedCourse(
                        source_row=row,
                        student_name=student_name,
                        course_type=course_type,
                        teacher_name=teacher_name,
                        course_date=f"{year}-{month:02d}-{day:02d}",
                        start_time=start_time,
                        end_time=end_time,
                        duration_hours=tail_duration or computed_duration,
                    )
                )
                continue

            subject = subject_from_text(line)
            if subject:
                current_subject = subject
                rest = SUBJECT_LINE_RE.sub("", line)
                inline_teacher = normalize_teacher_name(strip_hours(rest))
                if inline_teacher and inline_teacher in known_teachers:
                    current_teacher = inline_teacher
                elif current_subject in teacher_map:
                    current_teacher = teacher_map[current_subject]
                elif current_subject == "基础":
                    current_teacher = teacher_map.get("口语") or current_teacher
                continue

            standalone_teacher = normalize_teacher_name(line)
            if standalone_teacher and standalone_teacher in known_teachers:
                current_teacher = standalone_teacher

    return students, known_teachers, courses


def get_or_create_student(cursor: sqlite3.Cursor, student: dict) -> tuple[int, bool]:
    existing = cursor.execute("SELECT id, note FROM students WHERE name = ?", (student["name"],)).fetchone()
    if existing:
        student_id, existing_note = existing
        if student["note"] not in (existing_note or ""):
            merged_note = f"{existing_note or ''}\n{student['note']}".strip()
            cursor.execute(
                "UPDATE students SET status = ?, note = ? WHERE id = ?",
                (student["status"], merged_note, student_id),
            )
        return student_id, False

    cursor.execute(
        "INSERT INTO students (name, phone, status, note, created_at) VALUES (?, ?, ?, ?, ?)",
        (student["name"], None, student["status"], student["note"], now_iso()),
    )
    return cursor.lastrowid, True


def get_or_create_teacher(cursor: sqlite3.Cursor, name: str, subject: str) -> tuple[int, bool]:
    existing = cursor.execute("SELECT id FROM teachers WHERE name = ?", (name,)).fetchone()
    if existing:
        return existing[0], False

    cursor.execute(
        "INSERT INTO teachers (name, subject, phone, note, created_at) VALUES (?, ?, ?, ?, ?)",
        (name, subject, None, "来源：7月课时.xlsx 自动导入", now_iso()),
    )
    return cursor.lastrowid, True


def course_exists(cursor: sqlite3.Cursor, student_id: int, teacher_id: int, course: ParsedCourse) -> bool:
    existing = cursor.execute(
        """
        SELECT id FROM courses
        WHERE student_id = ?
          AND teacher_id = ?
          AND course_type = ?
          AND course_date = ?
          AND start_time = ?
          AND end_time = ?
        """,
        (
            student_id,
            teacher_id,
            course.course_type,
            course.course_date,
            course.start_time,
            course.end_time,
        ),
    ).fetchone()
    return existing is not None


def import_data(db_path: Path, students: list[dict], courses: list[ParsedCourse]) -> dict:
    backup_path = db_path.with_name(f"{db_path.stem}_before_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    if db_path.exists():
        shutil.copy2(db_path, backup_path)

    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        student_ids: dict[str, int] = {}
        teacher_ids: dict[str, int] = {}
        created_students = 0
        updated_students = 0
        created_teachers = 0
        inserted_courses = 0
        skipped_courses = 0

        for student in students:
            student_id, created = get_or_create_student(cursor, student)
            student_ids[student["name"]] = student_id
            if created:
                created_students += 1
            else:
                updated_students += 1

        for course in courses:
            teacher_id, created = get_or_create_teacher(cursor, course.teacher_name, course.course_type)
            teacher_ids[course.teacher_name] = teacher_id
            if created:
                created_teachers += 1

            student_id = student_ids[course.student_name]
            if course_exists(cursor, student_id, teacher_id, course):
                skipped_courses += 1
                continue

            now = now_iso()
            cursor.execute(
                """
                INSERT INTO courses (
                    student_id, teacher_id, course_type, course_date, start_time, end_time,
                    duration_hours, location, note, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    student_id,
                    teacher_id,
                    course.course_type,
                    course.course_date,
                    course.start_time,
                    course.end_time,
                    course.duration_hours,
                    None,
                    f"来源：7月课时.xlsx 第 {course.source_row} 行",
                    now,
                    now,
                ),
            )
            inserted_courses += 1

        conn.commit()
        return {
            "backup_path": str(backup_path) if db_path.exists() else "",
            "created_students": created_students,
            "updated_students": updated_students,
            "created_teachers": created_teachers,
            "inserted_courses": inserted_courses,
            "skipped_courses": skipped_courses,
            "unknown_teacher_courses": sum(1 for course in courses if course.teacher_name == "未分配老师"),
        }
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import July schedule data from Excel into SQLite.")
    parser.add_argument("--xlsx", default="/Users/sxw/工作/7月课时.xlsx")
    parser.add_argument("--db", default=str(Path(__file__).resolve().parents[1] / "ielts_scheduler.db"))
    parser.add_argument("--sheet", default="25年7月")
    parser.add_argument("--year", type=int, default=2025)
    parser.add_argument("--month", type=int, default=7)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    students, known_teachers, courses = parse_workbook(Path(args.xlsx), args.sheet, args.year, args.month)
    distinct_students = {student["name"] for student in students}

    print(f"解析学生：{len(distinct_students)}")
    print(f"解析老师：{len(known_teachers)}")
    print(f"解析 {args.year}-{args.month:02d} 课程：{len(courses)}")
    print(f"未识别老师课程：{sum(1 for course in courses if course.teacher_name == '未分配老师')}")

    if args.dry_run:
        print("dry-run 模式：未写入数据库。")
        return

    result = import_data(Path(args.db), students, courses)
    print("导入完成：")
    for key, value in result.items():
        print(f"- {key}: {value}")


if __name__ == "__main__":
    main()
